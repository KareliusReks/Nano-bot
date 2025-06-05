import os
import sqlite3
import pandas as pd
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler, ConversationHandler,
    ContextTypes, filters, CallbackContext
)
import openpyxl
from openpyxl.utils import get_column_letter

TOKEN = "7937094483:AAHi6-OUns9XstVRomhvIbaXDOpvCkhJ_Po"
DB_FILE = "users.db"
EXCEL_FILE = "данные.xlsx"
ADMIN_EXCEL_FILE = "админы.xlsx"

SUPERADMIN_LOGIN = "kudratullomuradov93@gmail.com"
SUPERADMIN_PASS = "KareliusReks"
SUPERADMIN_FIO = "Мурадов Кудратулло Абдурахмоновович"

(
    LOGIN, PASS, MENU,
    ADD_USER_FIO, ADD_USER_LOGIN, ADD_USER_PASS,
    ADD_ADMIN_FIO, ADD_ADMIN_LOGIN, ADD_ADMIN_PASS,
    FIND_FIO, DEL_FIO,
    DEL_ADMIN_LOGIN, DEL_ADMIN_CONFIRM
) = range(13)

sessions = {}

def ensure_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS admins (
            user_id INTEGER PRIMARY KEY,
            ФИО TEXT NOT NULL,
            Логин TEXT NOT NULL UNIQUE,
            Пароль TEXT NOT NULL
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            ФИО TEXT PRIMARY KEY,
            Логин TEXT NOT NULL,
            Пароль TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()

def log_action(user_id, action):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fio = sessions.get(user_id, {}).get('fio', 'Неизвестно')
    with open("admin_log.txt", "a", encoding="utf-8") as f:
        f.write(f"[{now}] {fio} ({user_id}): {action}\n")

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Введите логин администратора:")
    return LOGIN

async def get_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    login = update.message.text.strip()
    user_id = update.effective_user.id
    context.user_data['login'] = login

    if login == SUPERADMIN_LOGIN:
        context.user_data['fio'] = SUPERADMIN_FIO
        context.user_data['admin_pass'] = SUPERADMIN_PASS
        await update.message.reply_text("Введите пароль суперадмина:")
        return PASS

    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT ФИО, Пароль FROM admins WHERE Логин = ?", (login,))
    row = cursor.fetchone()
    conn.close()
    if row:
        context.user_data['fio'] = row[0]
        context.user_data['admin_pass'] = row[1]
        await update.message.reply_text("Введите пароль администратора:")
        return PASS
    else:
        await update.message.reply_text("❌ Вы не зарегистрированы. Доступ запрещён.")
        return ConversationHandler.END

async def check_admin_pass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    password = update.message.text.strip()
    user_id = update.effective_user.id
    if password == context.user_data['admin_pass']:
        sessions[user_id] = {'fio': context.user_data['fio'], 'login': context.user_data['login']}
        log_action(user_id, "Вход администратора")
        await show_menu(update, context)
        return MENU
    else:
        await update.message.reply_text("❌ Неверный пароль. /start")
        return ConversationHandler.END

async def show_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    login = sessions[user_id]['login']
    if login == SUPERADMIN_LOGIN:
        keyboard = [
            ["Зарегистрировать админа"],
            ["Добавить пользователя", "Найти"],
            ["Удалить", "Удалить админа", "Выгрузить"],
            ["Выйти"],
            ["Удалить все", "Выгрузить админов"]
        ]
    else:
        keyboard = [
            ["Добавить пользователя", "Найти"],
            ["Удалить", "Выгрузить"],
            ["Выйти"]
        ]
    await update.message.reply_text("Выберите действие:", reply_markup=ReplyKeyboardMarkup(keyboard, resize_keyboard=True))

async def menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    text = update.message.text.strip().lower()
    login = sessions[user_id]['login']

    if login == SUPERADMIN_LOGIN and text == "зарегистрировать админа":
        await update.message.reply_text("Введите ФИО администратора:")
        return ADD_ADMIN_FIO
    if text == "добавить пользователя":
        await update.message.reply_text("Введите ФИО пользователя:")
        return ADD_USER_FIO
    if text == "найти":
        await update.message.reply_text("Введите ФИО для поиска:")
        return FIND_FIO
    if text == "удалить":
        await update.message.reply_text("Введите ФИО пользователя для удаления:")
        return DEL_FIO
    if login == SUPERADMIN_LOGIN and text == "удалить админа":
        await update.message.reply_text("Введите логин администратора для удаления:")
        return DEL_ADMIN_LOGIN
    if text == "выгрузить":
        return await export_excel(update, context)
    if login == SUPERADMIN_LOGIN and text == "выгрузить админов":
        return await export_admins(update, context)
    if login == SUPERADMIN_LOGIN and text == "удалить все":
        return await delete_all(update, context)
    if text == "выйти":
        return await logout(update, context)
    await update.message.reply_text("Неизвестная команда. Выберите действие через кнопки меню.")
    return MENU

# --- Добавление пользователя ---
async def add_user_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['user_fio'] = update.message.text.strip()
    await update.message.reply_text("Введите логин пользователя:")
    return ADD_USER_LOGIN

async def add_user_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['user_login'] = update.message.text.strip()
    await update.message.reply_text("Введите пароль пользователя:")
    return ADD_USER_PASS

async def add_user_pass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = context.user_data['user_fio']
    login = context.user_data['user_login']
    password = update.message.text.strip()
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("INSERT OR REPLACE INTO users (ФИО, Логин, Пароль) VALUES (?, ?, ?)", (fio, login, password))
    conn.commit()
    conn.close()
    log_action(update.effective_user.id, f"Добавлен пользователь: {fio}")
    await update.message.reply_text(f"✅ Пользователь {fio} добавлен.")
    await show_menu(update, context)
    return MENU

# --- Добавление админа ---
async def add_admin_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['admin_fio'] = update.message.text.strip()
    await update.message.reply_text("Введите логин администратора:")
    return ADD_ADMIN_LOGIN

async def add_admin_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['admin_login'] = update.message.text.strip()
    await update.message.reply_text("Введите пароль администратора:")
    return ADD_ADMIN_PASS

async def add_admin_pass(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = context.user_data['admin_fio']
    login = context.user_data['admin_login']
    password = update.message.text.strip()
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    user_id = int(datetime.now().timestamp())  # уникальный user_id
    cursor.execute("INSERT OR REPLACE INTO admins (user_id, ФИО, Логин, Пароль) VALUES (?, ?, ?, ?)", (user_id, fio, login, password))
    conn.commit()
    conn.close()
    log_action(update.effective_user.id, f"Зарегистрирован админ: {fio}")
    await update.message.reply_text(f"✅ Администратор {fio} добавлен.")
    await show_menu(update, context)
    return MENU

# --- Удаление админа суперадмином ---
async def del_admin_login(update: Update, context: ContextTypes.DEFAULT_TYPE):
    login = update.message.text.strip()
    if login == SUPERADMIN_LOGIN:
        await update.message.reply_text("❌ Суперадмин не может удалить себя.")
        return await show_menu(update, context)
    context.user_data['del_admin_login'] = login
    await update.message.reply_text(f"Вы уверены, что хотите удалить администратора с логином {login}? (да/нет)")
    return DEL_ADMIN_CONFIRM

async def del_admin_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    answer = update.message.text.lower()
    if answer == "да":
        login = context.user_data.get('del_admin_login')
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM admins WHERE Логин = ?", (login,))
        conn.commit()
        conn.close()
        log_action(update.effective_user.id, f"Удалён админ: {login}")
        await update.message.reply_text(f"✅ Администратор {login} удалён.")
    else:
        await update.message.reply_text("❌ Отмена удаления администратора.")
    await show_menu(update, context)
    return MENU

# --- Поиск пользователя ---
async def find_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = update.message.text.strip()
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT Логин, Пароль FROM users WHERE ФИО = ?", (fio,))
    row = cursor.fetchone()
    conn.close()
    if row:
        await update.message.reply_text(f"👤 Логин: {row[0]}\n🔐 Пароль: {row[1]}")
    else:
        await update.message.reply_text("❌ Пользователь не найден.")
    await show_menu(update, context)
    return MENU

# --- Выгрузка пользователей в Excel с автошириной ---
async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT * FROM users", conn)
    conn.close()
    df.to_excel(EXCEL_FILE, index=False)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length + 3, 15)
    wb.save(EXCEL_FILE)
    await update.message.reply_document(open(EXCEL_FILE, "rb"))
    log_action(update.effective_user.id, "Экспорт пользователей")
    await show_menu(update, context)
    return MENU

# --- Выгрузка админов в Excel с автошириной ---
async def export_admins(update: Update, context: ContextTypes.DEFAULT_TYPE):
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    df = pd.read_sql_query("SELECT * FROM admins", conn)
    conn.close()
    df.to_excel(ADMIN_EXCEL_FILE, index=False)
    wb = openpyxl.load_workbook(ADMIN_EXCEL_FILE)
    ws = wb.active
    for col in ws.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_length + 3, 15)
    wb.save(ADMIN_EXCEL_FILE)
    await update.message.reply_document(open(ADMIN_EXCEL_FILE, "rb"))
    log_action(update.effective_user.id, "Экспорт админов")
    await show_menu(update, context)
    return MENU

# --- Удаление пользователя ---
async def del_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    fio = update.message.text.strip()
    if not fio or fio.lower() in [
        "добавить пользователя", "найти", "удалить", "выгрузить",
        "зарегистрировать админа", "удалить админа", "выгрузить админов",
        "удалить все", "выйти"
    ]:
        await update.message.reply_text("Введите корректное ФИО пользователя для удаления:")
        return DEL_FIO
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE ФИО = ?", (fio,))
    if cursor.fetchone():
        cursor.execute("DELETE FROM users WHERE ФИО = ?", (fio,))
        conn.commit()
        await update.message.reply_text(f"✅ Пользователь {fio} удалён.")
        log_action(update.effective_user.id, f"Удалён пользователь: {fio}")
    else:
        await update.message.reply_text("❌ Пользователь не найден.")
    conn.close()
    await show_menu(update, context)
    return MENU

# --- Удаление всех пользователей ---
async def delete_all(update: Update, context: ContextTypes.DEFAULT_TYPE):
    login = sessions[update.effective_user.id]['login']
    if login != SUPERADMIN_LOGIN:
        await update.message.reply_text("❌ Только суперадмин может удалять всех пользователей.")
        return await show_menu(update, context)
    ensure_db()
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM users")
    conn.commit()
    conn.close()
    log_action(update.effective_user.id, "Удалены все пользователи")
    await update.message.reply_text("🗑 Все пользователи удалены.")
    await show_menu(update, context)
    return MENU

# --- Выход ---
async def logout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sessions.pop(user_id, None)
    await update.message.reply_text("✅ Вы вышли из системы.", reply_markup=ReplyKeyboardRemove())
    return ConversationHandler.END

# --- MAIN ---
def main():
    ensure_db()
    application = Application.builder().token(TOKEN).build()
    conv = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_login)],
            PASS: [MessageHandler(filters.TEXT & ~filters.COMMAND, check_admin_pass)],
            MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, menu_handler)],
            ADD_USER_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_user_fio)],
            ADD_USER_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_user_login)],
            ADD_USER_PASS: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_user_pass)],
            ADD_ADMIN_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_admin_fio)],
            ADD_ADMIN_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_admin_login)],
            ADD_ADMIN_PASS: [MessageHandler(filters.TEXT & ~filters.COMMAND, add_admin_pass)],
            FIND_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, find_fio)],
            DEL_FIO: [MessageHandler(filters.TEXT & ~filters.COMMAND, del_fio)],
            DEL_ADMIN_LOGIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, del_admin_login)],
            DEL_ADMIN_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, del_admin_confirm)],
        },
        fallbacks=[],
        allow_reentry=True
    )
    application.add_handler(conv)
    application.add_handler(CommandHandler("logout", logout))
    application.run_polling()

if __name__ == "__main__":
    main()
